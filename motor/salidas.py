"""Salidas: ficheros de resultado generados a partir de un SELECT.

Una salida es un `SELECT` (sobre cualquier tabla o vista del almacén) volcado
a un fichero cuyo nombre se compone con la fecha y datos de la ejecución, por
ejemplo `20260805_previ_ok.xlsx`.

Se declaran en la definición de carga y se generan al terminar una carga
correcta, o se piden a mano con `etl salida`. A diferencia de
`etl exportar <vista>`, que vuelca una vista de consumo a un nombre fijo en
parquet y CSV, aquí el SQL y el nombre son libres.
"""

from datetime import datetime
from pathlib import Path

from . import entorno

ROOT = Path(__file__).resolve().parent.parent
EXPORT_DIR = entorno.ruta("export")

FORMATOS = {".csv": "csv", ".parquet": "parquet", ".xlsx": "xlsx"}

# Un color en hexadecimal, con o sin almohadilla: "000000", "#1F4E79".
HEX = "^#?[0-9A-Fa-f]{6}$"

# Presentación de una salida xlsx. Es deliberadamente pobre: cubre lo que hace
# que un fichero se pueda mandar a alguien sin retocarlo a mano (dónde empieza
# la tabla, cabecera distinguible, bordes, anchos y cómo se ven fechas e
# importes) y nada más. Un xlsx no es un informe: si hace falta algo que esto
# no da, el sitio es una plantilla, no más claves aquí.
SCHEMA_ESTILO = {
    "type": "object",
    "properties": {
        # Dónde cae la esquina de la cabecera. Dejar un margen arriba y a la
        # izquierda es lo que separa una tabla volcada de una tabla presentada.
        "fila_inicio": {"type": "integer", "minimum": 1},
        "columna_inicio": {"type": "integer", "minimum": 1},
        # Ancho de las columnas de margen que quedan a la izquierda.
        "ancho_margen": {"type": "number", "exclusiveMinimum": 0},
        "hoja": {"type": "string", "minLength": 1},
        "cabecera": {
            "type": "object",
            "properties": {
                "fondo": {"type": "string", "pattern": HEX},
                "texto": {"type": "string", "pattern": HEX},
                "negrita": {"type": "boolean"},
            },
            "additionalProperties": False,
        },
        "bordes": {"enum": ["thin", "medium", "none"]},
        "autofiltro": {"type": "boolean"},
        # Por nombre de columna del SELECT, no por letra de Excel: la letra
        # depende de columna_inicio y se rompe al reordenar el SELECT.
        "anchos": {"type": "object", "additionalProperties": {"type": "number"}},
        "formatos": {"type": "object", "additionalProperties": {"type": "string"}},
    },
    "additionalProperties": False,
}

SCHEMA_SALIDA = {
    "type": "object",
    "properties": {
        "nombre": {"type": "string", "minLength": 1},
        "fichero": {"type": "string", "minLength": 1},
        "sql": {"type": "string", "minLength": 1},
        "carpeta": {"type": "string", "minLength": 1},
        "estilo": SCHEMA_ESTILO,
    },
    "required": ["nombre", "fichero", "sql"],
    "additionalProperties": False,
}


def formato_de(nombre_fichero: str) -> str:
    sufijo = Path(nombre_fichero).suffix.lower()
    if sufijo not in FORMATOS:
        raise ValueError(
            f"extensión '{sufijo}' no soportada en salidas (usa {', '.join(sorted(FORMATOS))})"
        )
    return FORMATOS[sufijo]


def resolver_nombre(plantilla: str, contexto: dict, momento: datetime = None) -> str:
    """Compone el nombre del fichero.

    Admite marcas de fecha de strftime (`%Y%m%d` -> 20260805) y campos entre
    llaves del contexto de la ejecución (`{carga}`, `{ejecucion_id}`).
    """
    momento = momento or datetime.now()
    return momento.strftime(plantilla).format(**contexto)


def _escribir_xlsx(con, sql: str, ruta: Path, valores: list = None) -> None:
    """Escribe xlsx con la extensión `excel` de DuckDB, que resuelve el volcado
    dentro del motor. Si no está disponible (por ejemplo, sin red para
    instalarla la primera vez), se recurre a openpyxl, que ya es dependencia
    del proyecto pero pasa las filas por Python.

    `sql` llega ya con sus marcadores y `valores` con lo que enlazar."""
    valores = valores or []
    try:
        con.execute("INSTALL excel; LOAD excel;")
        # HEADER true es obligatorio aquí: sin él la primera fila del fichero ya
        # es un dato y quien lo abra en Excel no sabe qué columna es cada cosa.
        con.execute(
            f"COPY ({sql}) TO '{ruta.as_posix()}' (FORMAT xlsx, HEADER true)", valores
        )
        return
    except Exception:
        pass

    import openpyxl

    cursor = con.execute(sql, valores)
    columnas = [d[0] for d in cursor.description]
    wb = openpyxl.Workbook(write_only=True)
    ws = wb.create_sheet()
    ws.append(columnas)
    for fila in cursor.fetchall():
        ws.append(list(fila))
    wb.save(ruta)


def _argb(color: str) -> str:
    """Hexadecimal de usuario a ARGB de Excel: '#1F4E79' -> 'FF1F4E79'."""
    return "FF" + color.lstrip("#").upper()


def _escribir_xlsx_con_estilo(con, sql: str, ruta: Path, valores: list, estilo: dict) -> None:
    """Escribe xlsx con presentación, siempre por openpyxl.

    No hay camino por DuckDB aquí: su `COPY ... (FORMAT xlsx)` vuelca la
    rejilla y no expone estilos, así que una salida con `estilo` paga el paso
    por Python a cambio del formato. Es un intercambio consciente y solo lo
    paga quien lo pide: sin `estilo`, la salida sigue yendo por el motor.
    """
    import openpyxl
    from openpyxl.styles import Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    cursor = con.execute(sql, valores)
    columnas = [d[0] for d in cursor.description]
    filas = cursor.fetchall()

    fila_cab = estilo.get("fila_inicio", 1)
    col_ini = estilo.get("columna_inicio", 1)

    wb = openpyxl.Workbook()
    ws = wb.active
    if estilo.get("hoja"):
        ws.title = estilo["hoja"]

    for desplazamiento, nombre in enumerate(columnas):
        ws.cell(row=fila_cab, column=col_ini + desplazamiento, value=nombre)
    for numero, fila in enumerate(filas, start=fila_cab + 1):
        for desplazamiento, valor in enumerate(fila):
            ws.cell(row=numero, column=col_ini + desplazamiento, value=valor)

    fila_final = fila_cab + len(filas)
    col_final = col_ini + len(columnas) - 1

    cabecera = estilo.get("cabecera") or {}
    if cabecera:
        relleno = PatternFill("solid", fgColor=_argb(cabecera["fondo"])) if cabecera.get("fondo") else None
        fuente = Font(
            color=_argb(cabecera["texto"]) if cabecera.get("texto") else None,
            bold=cabecera.get("negrita", False),
        )
        for numero in range(col_ini, col_final + 1):
            celda = ws.cell(row=fila_cab, column=numero)
            if relleno is not None:
                celda.fill = relleno
            celda.font = fuente

    estilo_borde = estilo.get("bordes", "none")
    if estilo_borde != "none":
        lado = Side(style=estilo_borde, color="FF000000")
        borde = Border(left=lado, right=lado, top=lado, bottom=lado)
        for fila_celdas in ws.iter_rows(min_row=fila_cab, max_row=fila_final,
                                        min_col=col_ini, max_col=col_final):
            for celda in fila_celdas:
                celda.border = borde

    # Los formatos van por nombre de columna del SELECT y solo a los datos: la
    # cabecera es texto y un '#,##0.00' encima de "Importe" no pinta nada.
    for nombre, formato in (estilo.get("formatos") or {}).items():
        if nombre not in columnas:
            raise ValueError(
                f"la salida da formato a '{nombre}', que no es ninguna columna del SELECT "
                f"({', '.join(columnas)})"
            )
        columna = col_ini + columnas.index(nombre)
        for numero in range(fila_cab + 1, fila_final + 1):
            ws.cell(row=numero, column=columna).number_format = formato

    for nombre, ancho in (estilo.get("anchos") or {}).items():
        if nombre not in columnas:
            raise ValueError(
                f"la salida da ancho a '{nombre}', que no es ninguna columna del SELECT "
                f"({', '.join(columnas)})"
            )
        letra = get_column_letter(col_ini + columnas.index(nombre))
        ws.column_dimensions[letra].width = ancho

    if estilo.get("ancho_margen"):
        for numero in range(1, col_ini):
            ws.column_dimensions[get_column_letter(numero)].width = estilo["ancho_margen"]

    if estilo.get("autofiltro"):
        ws.auto_filter.ref = (
            f"{get_column_letter(col_ini)}{fila_cab}:{get_column_letter(col_final)}{fila_final}"
        )

    wb.save(ruta)


def generar(con, salida: dict, contexto: dict = None) -> dict:
    """Genera una salida. Devuelve la ruta y el número de filas escritas.

    El mismo `contexto` sirve para las dos cosas: las llaves del nombre de
    fichero (`{carga}`) y las variables del SQL (`$p_tienda`). Son sintaxis
    distintas a propósito, porque el nombre de fichero se compone pegando
    texto y el SQL nunca.
    """
    from . import sustitucion

    contexto = contexto or {}
    nombre_fichero = resolver_nombre(salida["fichero"], contexto)
    formato = formato_de(nombre_fichero)

    carpeta = Path(salida.get("carpeta", EXPORT_DIR))
    if not carpeta.is_absolute():
        carpeta = ROOT / carpeta
    carpeta.mkdir(parents=True, exist_ok=True)
    ruta = carpeta / nombre_fichero

    # Se resuelve una vez y se reutiliza: el SQL entra en tres sentencias
    # distintas (contar, copiar, volcar) y las tres necesitan los mismos
    # valores enlazados en el mismo orden.
    sql, valores = sustitucion.resolver(salida["sql"], contexto)
    filas = con.execute(f"SELECT count(*) FROM ({sql})", valores).fetchone()[0]

    if formato == "csv":
        con.execute(f"COPY ({sql}) TO '{ruta.as_posix()}' (FORMAT CSV, HEADER)", valores)
    elif formato == "parquet":
        con.execute(f"COPY ({sql}) TO '{ruta.as_posix()}' (FORMAT PARQUET)", valores)
    elif salida.get("estilo"):
        _escribir_xlsx_con_estilo(con, sql, ruta, valores, salida["estilo"])
    else:
        _escribir_xlsx(con, sql, ruta, valores)

    return {"nombre": salida["nombre"], "fichero": str(ruta), "filas": filas, "formato": formato}


def generar_todas(con, definicion: dict, contexto: dict = None) -> list:
    return [generar(con, salida, contexto) for salida in definicion.get("salidas", [])]
