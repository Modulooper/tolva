"""Orquestador del motor ETL: escaneo, mapping fila a fila, promoción, rechazos.

Ningún registro pasa por el modelo. La definición de carga (ya revisada por
el usuario) es lo único que decide cómo se transforma cada columna.

Las cargas de fichero no hacen upsert fila a fila: promueven en bloque
(borrado de las combinaciones de `campos_singularidad` presentes en los datos
nuevos + inserción masiva). Ver `motor/cargas.py` para las dos formas de
carga (directa y con tabla hall).
"""

import csv
import getpass
import json
import time
from dataclasses import dataclass
from pathlib import Path

from . import cargas, db, documentos, operaciones, parametros, salidas, sustitucion, validaciones


class VariableInvalidaError(ValueError):
    """El SQL de una variable no se pudo evaluar o no devolvió una sola fila.
    Es un fallo de la definición, no un dato malo: se trata como error duro
    para no seguir con una variable que no vale lo que dice."""


# Una sola implementación del hash para todo el sistema: el que identifica la
# ejecución es el mismo que direcciona el documento en el almacén.
_hash_fichero = documentos.hash_fichero


def _leer_csv(ruta: Path, delimitador: str, fila_cabecera: int, encoding: str):
    with ruta.open(encoding=encoding, newline="") as f:
        filas = list(csv.reader(f, delimiter=delimitador))
    cabecera = filas[fila_cabecera - 1]
    datos = filas[fila_cabecera:]
    return cabecera, [dict(zip(cabecera, fila)) for fila in datos]


def _leer_excel_openpyxl(ruta: Path, hoja, fila_cabecera: int):
    import openpyxl

    wb = openpyxl.load_workbook(ruta, data_only=True, read_only=True)
    if isinstance(hoja, str) and hoja:
        ws = wb[hoja]
    elif isinstance(hoja, int):
        ws = wb.worksheets[hoja]
    else:
        ws = wb.active
    filas = list(ws.iter_rows(values_only=True))
    cabecera = [str(c) if c is not None else "" for c in filas[fila_cabecera - 1]]
    datos = filas[fila_cabecera:]
    return cabecera, [dict(zip(cabecera, fila)) for fila in datos]


def _leer_excel_duckdb(ruta: Path, hoja: str):
    """Lee el xlsx con la extensión `excel` de DuckDB.

    Medido sobre un fichero de 44,8 MB y 497.383 filas: 8 s frente a los 71 s
    de openpyxl, que parsea el XML en Python.

    `all_varchar=true` deja todos los valores como texto, igual que hace el
    lector de CSV, para que sea el mapping declarado —y no el lector— quien
    decida los tipos. Efecto secundario deseable: una celda con fecha real de
    Excel llega como serial ("46101") en vez de como `datetime`, y el serial sí
    lo resuelve `motor/fechas.py`; con openpyxl esa misma celda hacía fallar el
    parseo y la fila acababa rechazada.
    """
    import duckdb

    con = duckdb.connect(":memory:")
    try:
        con.execute("INSTALL excel; LOAD excel;")
        consulta = f"SELECT * FROM read_xlsx('{ruta.as_posix()}', sheet=?, all_varchar=true)"
        cursor = con.execute(consulta, [hoja])
        cabecera = [d[0] for d in cursor.description]
        filas = cursor.fetchall()
    finally:
        con.close()
    return cabecera, [dict(zip(cabecera, fila)) for fila in filas]


def _leer_excel(ruta: Path, hoja, fila_cabecera: int):
    # El lector de DuckDB toma la primera fila como cabecera y necesita el
    # nombre de la hoja; fuera de ese caso, y si la extensión no está
    # disponible (requiere red la primera vez), se usa openpyxl.
    if fila_cabecera == 1 and isinstance(hoja, str) and hoja:
        try:
            return _leer_excel_duckdb(ruta, hoja)
        except Exception:
            pass
    return _leer_excel_openpyxl(ruta, hoja, fila_cabecera)


def celdas_referenciadas(definicion: dict) -> list:
    """Las celdas que pide el mapping, en orden y sin repetir."""
    refs = []
    for campo in definicion.get("mapping", []):
        for op in campo.get("operaciones", []):
            if op.get("tipo") == "celda" and op["referencia"] not in refs:
                refs.append(op["referencia"])
    return refs


def leer_celdas(ruta: Path, definicion: dict) -> dict:
    """{referencia: valor} de las celdas sueltas que pide el mapping.

    Se abre el libro una segunda vez, con openpyxl, porque el lector rápido de
    DuckDB devuelve una tabla y aquí hacen falta posiciones concretas. Solo
    ocurre si la carga usa `celda`, así que quien no las use no lo paga.
    """
    refs = celdas_referenciadas(definicion)
    if not refs:
        return {}

    import openpyxl

    wb = openpyxl.load_workbook(ruta, data_only=True, read_only=False)
    try:
        hoja = definicion.get("hoja")
        if isinstance(hoja, str) and hoja:
            ws = wb[hoja]
        elif isinstance(hoja, int):
            ws = wb.worksheets[hoja]
        else:
            ws = wb.active
        return {ref: ws[ref].value for ref in refs}
    finally:
        wb.close()


def leer_fichero(ruta: Path, definicion: dict):
    if definicion["formato"] == "csv":
        return _leer_csv(
            ruta,
            definicion.get("delimitador", ","),
            definicion["fila_cabecera"],
            definicion.get("encoding", "utf-8-sig"),
        )
    return _leer_excel(ruta, definicion.get("hoja"), definicion["fila_cabecera"])


@dataclass
class ResultadoProceso:
    filas_leidas: int
    validas: list
    rechazadas: list  # tuplas (num_fila, motivo, campo_implicado, contenido_raw)
    columnas_extra: set


def procesar_filas(cabecera: list, filas: list, definicion: dict, valores_parametros=None,
                   celdas=None) -> ResultadoProceso:
    mapping = definicion["mapping"]
    origenes_declarados = {c["origen"] for c in mapping if "origen" in c}
    columnas_extra = {c for c in cabecera if c and c not in origenes_declarados}
    valores_parametros = valores_parametros or {}
    celdas = celdas or {}

    contextos = {}
    for campo in mapping:
        if "origen" not in campo:
            contextos[campo["destino"]] = {"parametros": valores_parametros, "celdas": celdas}
            continue
        valores_columna = [fila.get(campo["origen"]) for fila in filas]
        contexto = operaciones.preparar_contexto(campo, valores_columna)
        contexto["parametros"] = valores_parametros
        contexto["celdas"] = celdas
        contextos[campo["destino"]] = contexto

    validas, rechazadas = [], []
    for i, fila in enumerate(filas, start=definicion["fila_cabecera"] + 1):
        registro = {}
        error = None
        for campo in mapping:
            valor_origen = fila.get(campo["origen"]) if "origen" in campo else None
            try:
                registro[campo["destino"]] = operaciones.aplicar_cadena(
                    valor_origen, campo.get("operaciones", []), contextos[campo["destino"]]
                )
            except Exception as exc:
                error = (str(exc), campo["destino"])
                break
        if error:
            motivo, campo_implicado = error
            rechazadas.append(
                (i, motivo, campo_implicado, json.dumps(fila, default=str, ensure_ascii=False))
            )
            continue
        if columnas_extra:
            registro.setdefault(
                "extra_fields",
                json.dumps({c: fila.get(c) for c in columnas_extra}, default=str, ensure_ascii=False),
            )
        validas.append(registro)

    return ResultadoProceso(len(filas), validas, rechazadas, columnas_extra)


def _insertar_bloque(con, tabla: str, filas: list) -> None:
    """Inserción masiva vía DataFrame.

    DuckDB es columnar: una sentencia INSERT por fila (o una sola sentencia con
    miles de tuplas) es un antipatrón que degrada de forma no lineal — medido:
    1.000 filas por SQL tardan ~13s, mientras que 500.000 vía DataFrame tardan
    ~0,35s. Todo lo que escriba volumen en el almacén pasa por aquí.
    """
    if not filas:
        return
    import pandas as pd

    columnas = sorted({c for fila in filas for c in fila.keys()})
    df = pd.DataFrame(filas, columns=columnas)
    con.register("_bloque_entrante", df)
    try:
        con.execute(f"INSERT INTO {tabla} ({', '.join(columnas)}) SELECT * FROM _bloque_entrante")
    finally:
        con.unregister("_bloque_entrante")


TABLA_ENTRANTE = "_entrante"


def _columnas_de(con, tabla: str) -> set:
    filas = con.execute(
        "SELECT column_name FROM information_schema.columns WHERE table_name = ?", [tabla]
    ).fetchall()
    return {f[0] for f in filas}


def _crear_entrante(con, filas: list, ejecucion_id: int, columnas_mapping: list) -> None:
    """Materializa las filas ya mapeadas en una tabla temporal.

    Existe para que las validaciones puedan interrogar los datos entrantes con
    SQL antes de que toquen ninguna tabla real, también en cargas sin hall, y
    para que la promoción no tenga que volver a pasar por Python.
    """
    import pandas as pd

    columnas = sorted(set(columnas_mapping) | {c for fila in filas for c in fila.keys()})
    df = pd.DataFrame(filas, columns=columnas) if filas else pd.DataFrame(columns=columnas)
    df["ejecucion_id"] = ejecucion_id
    con.register("_df_entrante", df)
    try:
        con.execute(f"CREATE OR REPLACE TEMP TABLE {TABLA_ENTRANTE} AS SELECT * FROM _df_entrante")
    finally:
        con.unregister("_df_entrante")


def _cargar_hall(con, tabla_hall: str) -> None:
    """La hall es siempre foto completa: se vacía entera y se recarga desde
    los datos entrantes."""
    con.execute(f"DELETE FROM {tabla_hall}")
    comunes = sorted(_columnas_de(con, tabla_hall) & _columnas_de(con, TABLA_ENTRANTE))
    lista = ", ".join(comunes)
    con.execute(f"INSERT INTO {tabla_hall} ({lista}) SELECT {lista} FROM {TABLA_ENTRANTE}")


def _promover_desde_sql(con, tabla_destino: str, campos_singularidad: list, sql_origen: str,
                        contexto: dict = None) -> tuple:
    """Promoción sin pasar los datos por Python: el borrado por singularidad y
    la inserción se resuelven dentro del motor.

    Solo se promueven las columnas que existen en el destino: el origen puede
    traer columnas de trabajo (o `ejecucion_id` en tablas que no la tengan) que
    no deben viajar.
    """
    consulta, valores = sustitucion.resolver(sql_origen, contexto or {})
    con.execute(f"CREATE OR REPLACE TEMP TABLE _transformadas AS {consulta}", valores)
    promovidas = con.execute("SELECT count(*) FROM _transformadas").fetchone()[0]
    columnas = sorted(_columnas_de(con, "_transformadas") & _columnas_de(con, tabla_destino))
    lista = ", ".join(columnas)

    borradas = 0
    if campos_singularidad:
        condicion = " AND ".join(f"t.{c} IS NOT DISTINCT FROM e.{c}" for c in campos_singularidad)
        existe = f"EXISTS (SELECT 1 FROM _transformadas e WHERE {condicion})"
        borradas = con.execute(f"SELECT count(*) FROM {tabla_destino} t WHERE {existe}").fetchone()[0]
        con.execute(f"DELETE FROM {tabla_destino} t WHERE {existe}")

    con.execute(f"INSERT INTO {tabla_destino} ({lista}) SELECT {lista} FROM _transformadas")
    con.execute("DROP TABLE _transformadas")
    return promovidas, borradas


def _ejecutar_acciones(con, definicion: dict, momento: str, contexto: dict = None) -> list:
    """Acciones SQL declaradas para un momento del ciclo de vida."""
    ejecutadas = []
    for accion in definicion.get("acciones", []):
        if accion["momento"] == momento:
            sustitucion.ejecutar(con, accion["sql"], contexto or {})
            ejecutadas.append(accion["sql"])
    return ejecutadas


def _capturar_variables(con, definicion: dict, momento: str, contexto: dict) -> dict:
    """Evalúa las variables declaradas para este momento y las añade al
    contexto con prefijo `v_`.

    Cada columna de la fila resultante es una variable, así que una consulta
    puede definir varias de golpe. Que devuelva cero filas o más de una es un
    error duro y no un nulo silencioso: una variable vacía porque un WHERE no
    casó acaba en un UPDATE y el fallo se descubre en los datos, no aquí.
    """
    capturadas = {}
    for variable in definicion.get("variables", []):
        if variable.get("momento", "tras_validar") != momento:
            continue
        try:
            cursor = sustitucion.ejecutar(con, variable["sql"], contexto)
            columnas = [d[0] for d in cursor.description] if cursor.description else []
            filas = cursor.fetchall()
        except Exception as exc:
            raise VariableInvalidaError(
                f"la variable de momento '{momento}' no se pudo evaluar: {exc}"
            ) from exc
        if len(filas) != 1:
            raise VariableInvalidaError(
                f"el SQL de una variable ({momento}) devolvió {len(filas)} filas y "
                f"debe devolver exactamente 1: {variable['sql']}"
            )
        for columna, valor in zip(columnas, filas[0]):
            capturadas[columna] = valor
            contexto[f"v_{columna}"] = valor
    return capturadas


def _insertar_rechazos(con, ejecucion_id: int, rechazadas: list) -> None:
    _insertar_bloque(
        con,
        "_rechazos",
        [
            {
                "ejecucion_id": ejecucion_id,
                "num_fila": num_fila,
                "motivo": motivo,
                "campo_implicado": campo,
                "contenido_raw": raw,
            }
            for num_fila, motivo, campo, raw in rechazadas
        ],
    )


def _registrar_validaciones(con, ejecucion_id, origen: str, resultados: list) -> None:
    filas = [
        {
            "ejecucion_id": ejecucion_id,
            "origen": origen,
            "nombre": r["nombre"],
            "tipo": r["tipo"],
            "mensaje": r["mensaje"],
            "afectadas": r["afectadas"],
            "detalle": json.dumps(
                [dict(zip(r["columnas"], fila)) for fila in r["detalle"]],
                default=str,
                ensure_ascii=False,
            ),
        }
        for r in validaciones.disparadas(resultados)
    ]
    _insertar_bloque(con, "_validaciones_disparadas", filas)


def _validar_o_lanzar(definicion: dict, con) -> None:
    errores = cargas.validar(definicion, con)
    if errores:
        raise ValueError("definición inválida: " + "; ".join(errores))


def dry_run_carga(nombre_carga: str, db_path=None, valores_parametros=None) -> dict:
    """Ejecuta sin escribir nada en el almacén. Devuelve resultado y rechazos."""
    definicion = cargas.cargar(nombre_carga)
    con = db.conectar(db_path or db.DB_PATH)
    try:
        _validar_o_lanzar(definicion, con)
        resueltos = parametros.resolver(con, definicion, valores_parametros)
        carpeta = cargas.carpeta_entrada(definicion)
        ficheros = sorted(carpeta.glob(definicion["patron"]))
        resultados = []
        for ruta in ficheros:
            cabecera, filas = leer_fichero(ruta, definicion)
            # El dry-run tiene que ver exactamente lo que verá la carga: si no
            # leyera las celdas, esos campos saldrían vacíos aquí y con valor
            # al ejecutar, que es la peor forma de enterarse.
            resultado = procesar_filas(
                cabecera, filas, definicion, resueltos, leer_celdas(ruta, definicion)
            )
            resultados.append(
                {
                    "fichero": ruta.name,
                    "filas_leidas": resultado.filas_leidas,
                    "filas_ok": len(resultado.validas),
                    "filas_rechazadas": len(resultado.rechazadas),
                    "muestra_validas": resultado.validas[:5],
                    "muestra_rechazos": resultado.rechazadas[:5],
                    "columnas_extra": sorted(resultado.columnas_extra),
                }
            )
        return {"carga": definicion["nombre"], "ficheros": resultados}
    finally:
        con.close()


def _procesar_fichero(con, definicion: dict, ruta: Path, forzar: bool,
                      valores_parametros=None, registro_parametros=None) -> dict:
    hash_fichero = _hash_fichero(ruta)
    nombre_carga = definicion["nombre"]

    if not forzar:
        ya_ok = con.execute(
            "SELECT count(*) FROM _ejecuciones WHERE carga = ? AND hash_fichero = ? AND estado = 'OK'",
            [nombre_carga, hash_fichero],
        ).fetchone()[0]
        if ya_ok:
            return {"fichero": ruta.name, "estado": "OMITIDO", "motivo": "ya procesado (mismo hash)"}

    inicio = time.perf_counter()
    cabecera, filas = leer_fichero(ruta, definicion)
    celdas = leer_celdas(ruta, definicion)
    resultado = procesar_filas(cabecera, filas, definicion, valores_parametros, celdas)
    # Si se llega aquí, el fichero se leyó y procesó fila a fila sin excepción:
    # las filas inválidas quedan en _rechazos, no convierten la ejecución en un fallo.

    campos_singularidad = definicion.get("campos_singularidad", [])
    columnas_mapping = [c["destino"] for c in definicion["mapping"]]

    # La ejecución se registra ANTES de escribir para tener su id disponible:
    # las tablas que declaren `ejecucion_id` lo llevan en cada fila, y así se
    # puede borrar o inspeccionar exactamente lo que metió una carga concreta.
    # Además, una carga abortada por un stop deja traza (antes no dejaba
    # ninguna, porque el rollback se llevaba también el registro).
    con.execute("BEGIN TRANSACTION")
    try:
        ejecucion_id = con.execute(
            """INSERT INTO _ejecuciones
               (carga, tipo, fichero, hash_fichero, filas_leidas, filas_ok, filas_rechazadas,
                estado, usuario, parametros)
               VALUES (?, 'carga', ?, ?, ?, ?, ?, 'EN_CURSO', ?, ?) RETURNING id""",
            [
                nombre_carga,
                ruta.name,
                hash_fichero,
                resultado.filas_leidas,
                len(resultado.validas),
                len(resultado.rechazadas),
                getpass.getuser(),
                json.dumps(registro_parametros, default=str, ensure_ascii=False)
                if registro_parametros
                else None,
            ],
        ).fetchone()[0]
        # Una carga es siempre su propia principal (migración 013): el
        # historial encadenado es cosa de las ediciones del CLI.
        con.execute(
            "UPDATE _ejecuciones SET ejecucion_id_principal = id WHERE id = ?",
            [ejecucion_id],
        )

        # El fichero que originó la carga queda archivado y vinculado. Si la
        # carga acaba revertida, los bytes copiados quedan huérfanos en el
        # almacén, pero al estar direccionados por contenido no estorban: el
        # siguiente archivado del mismo fichero los reutiliza.
        documentos.archivar(con, ruta, ejecucion_id, documentos.TAG_ORIGEN)

        # El contexto de variables se construye aquí, con el id ya conocido, y
        # va creciendo a lo largo del ciclo: `$promovidas` no existe hasta que
        # hay promoción, y usarlo antes tiene que fallar diciéndolo.
        contexto = sustitucion.contexto_de(
            ejecucion_id=ejecucion_id,
            carga=nombre_carga,
            fichero=ruta.name,
            hash_fichero=hash_fichero,
            parametros=valores_parametros or {},
        )

        _ejecutar_acciones(con, definicion, "antes", contexto)
        _capturar_variables(con, definicion, "antes", contexto)
        _crear_entrante(con, resultado.validas, ejecucion_id, columnas_mapping)
        if cargas.usa_hall(definicion):
            _cargar_hall(con, definicion["tabla_hall"])

        resultados_validacion = validaciones.ejecutar(
            con, definicion.get("validaciones", []), contexto
        )
        _registrar_validaciones(con, ejecucion_id, f"carga:{nombre_carga}", resultados_validacion)

        if validaciones.hay_stop(resultados_validacion):
            # Un stop no revierte lo ya escrito en la hall: se conserva (y se
            # registra la ejecución) para poder investigar el fichero rechazado
            # con `db consultar`. Si se quiere limpiar, la carga declara una
            # acción con momento "al_fallar". El destino no se ha tocado.
            _ejecutar_acciones(con, definicion, "al_fallar", contexto)
            con.execute(
                "UPDATE _ejecuciones SET estado = 'ERROR', duracion = ? WHERE id = ?",
                [time.perf_counter() - inicio, ejecucion_id],
            )
            _insertar_rechazos(con, ejecucion_id, resultado.rechazadas)
            con.execute("COMMIT")
            return {
                "fichero": ruta.name,
                "estado": "ERROR",
                "ejecucion_id": ejecucion_id,
                "filas_leidas": resultado.filas_leidas,
                "filas_ok": len(resultado.validas),
                "filas_rechazadas": len(resultado.rechazadas),
                "filas_promovidas": 0,
                "filas_sustituidas": 0,
                "validaciones": resultados_validacion,
                "tabla_hall": definicion.get("tabla_hall"),
            }

        # Superados los stops, se ejecutan las acciones dependientes y se promueve.
        _capturar_variables(con, definicion, "tras_validar", contexto)
        _ejecutar_acciones(con, definicion, "tras_validar", contexto)
        origen = (
            definicion["transformacion_sql"]
            if cargas.usa_hall(definicion)
            else f"SELECT * FROM {TABLA_ENTRANTE}"
        )
        promovidas, borradas = _promover_desde_sql(
            con, definicion["tabla_destino"], campos_singularidad, origen, contexto
        )

        # Solo a partir de aquí existe el resultado de la escritura, y es lo
        # único que este momento aporta sobre `tras_validar`: derivar de lo que
        # de verdad quedó en el destino, sin reimplementar la singularidad.
        contexto["promovidas"] = promovidas
        contexto["borradas"] = borradas
        _capturar_variables(con, definicion, "tras_promover", contexto)
        _ejecutar_acciones(con, definicion, "tras_promover", contexto)

        con.execute(
            "UPDATE _ejecuciones SET estado = 'OK', duracion = ? WHERE id = ?",
            [time.perf_counter() - inicio, ejecucion_id],
        )
        _insertar_rechazos(con, ejecucion_id, resultado.rechazadas)
        con.execute("COMMIT")
    except Exception:
        con.execute("ROLLBACK")
        raise

    if resultado.columnas_extra:
        print(f"AVISO: columnas no declaradas en '{ruta.name}': {sorted(resultado.columnas_extra)}")

    # Las salidas se generan con la carga ya confirmada: escribir ficheros no
    # debe poder deshacer datos correctos, y el SELECT debe ver lo promovido.
    # Con el contexto completo: una salida puede filtrar por `$p_tienda` o por
    # `$v_total`, y nombrar el fichero con `{carga}` como siempre.
    ficheros_salida = salidas.generar_todas(con, definicion, contexto)

    return {
        "fichero": ruta.name,
        "estado": "OK",
        "ejecucion_id": ejecucion_id,
        "salidas": ficheros_salida,
        "filas_leidas": resultado.filas_leidas,
        "filas_ok": len(resultado.validas),
        "filas_rechazadas": len(resultado.rechazadas),
        "filas_promovidas": promovidas,
        "filas_sustituidas": borradas,
        "validaciones": resultados_validacion,
        "tabla_hall": definicion.get("tabla_hall"),
    }


def ejecutar_carga(nombre_carga: str, forzar: bool = False, db_path=None, valores_parametros=None) -> dict:
    definicion = cargas.cargar(nombre_carga)
    con = db.conectar(db_path or db.DB_PATH)
    try:
        _validar_o_lanzar(definicion, con)
        # Antes de leer nada: si falta un parámetro obligatorio, mejor fallar
        # aquí que a medio camino con la ejecución ya registrada.
        resueltos = parametros.resolver(con, definicion, valores_parametros)
        registro = parametros.registro(definicion, valores_parametros, resueltos)
        carpeta = cargas.carpeta_entrada(definicion)
        ficheros = sorted(carpeta.glob(definicion["patron"]))
        resumen = [
            _procesar_fichero(con, definicion, ruta, forzar, resueltos, registro)
            for ruta in ficheros
        ]
        return {
            "carga": definicion["nombre"],
            "ficheros": resumen,
            "estado": "ERROR" if any(f["estado"] == "ERROR" for f in resumen) else "OK",
        }
    finally:
        con.close()
