"""Variables de carga, momento `tras_promover` y operación `celda`.

Las tres cosas resuelven el mismo tipo de problema: un dato que la carga
necesita y que no está en ninguna columna del fichero. Antes había que
teclearlo al lanzar (parámetros) o deducirlo con SQL frágil.
"""

from motor import cargas, motor_etl
from pruebas.base import PruebaConAlmacen

DESCRIPCION = (
    "Carga de prueba para variables: no describe nada real, solo satisface el "
    "mínimo del esquema y da un sujeto sobre el que comprobar el ciclo."
)


class BaseVariables(PruebaConAlmacen):
    def setUp(self):
        super().setUp()
        self.con.execute("""
            CREATE TABLE destino (
                referencia VARCHAR, importe DOUBLE, origen_carga VARCHAR,
                extra_fields VARCHAR, ejecucion_id BIGINT
            )
        """)
        self.con.execute("CREATE TABLE auditoria (que VARCHAR, valor VARCHAR, ejecucion_id BIGINT)")
        self.escribir_catalogo(self.ficha_catalogo("destino", {
            "referencia": ("varchar", False), "importe": ("double", False),
            "origen_carga": ("varchar", False), "extra_fields": ("varchar", False),
            "ejecucion_id": ("integer", False),
        }))

    def carga(self, **extra):
        definicion = {
            "nombre": "ventas",
            "descripcion": DESCRIPCION,
            "patron": "*.csv",
            "formato": "csv",
            "delimitador": ";",
            "fila_cabecera": 1,
            "tabla_destino": "destino",
            "campos_singularidad": ["referencia"],
            "mapping": [
                {"origen": "Ref", "destino": "referencia", "operaciones": [{"tipo": "trim"}]},
                {"origen": "Importe", "destino": "importe",
                 "operaciones": [{"tipo": "cast", "tipo_destino": "double"}]},
            ],
        }
        definicion.update(extra)
        return self.escribir_carga(definicion)

    def csv(self, contenido="Ref;Importe\nA-1;10\nA-2;20\n"):
        self.escribir_csv("ventas", "ventas.csv", contenido)

    def ejecutar(self, nombre):
        return motor_etl.ejecutar_carga(nombre, db_path=self.db_path)


class PruebaVariablesDeSistema(BaseVariables):
    def test_ejecucion_id_es_la_de_verdad_no_un_max(self):
        """El motor conoce el id; no hay que deducirlo con `SELECT max(id)`,
        que acierta por casualidad y deja de acertar sin avisar."""
        nombre = self.carga(acciones=[{
            "momento": "tras_promover",
            "sql": "INSERT INTO auditoria VALUES ('ejecucion', NULL, $ejecucion_id)",
        }])
        self.csv()
        self.ejecutar(nombre)

        anotado = self.escalar("SELECT ejecucion_id FROM auditoria")
        real = self.escalar("SELECT id FROM _ejecuciones WHERE carga = 'ventas'")
        self.assertEqual(anotado, real)

    def test_carga_fichero_y_hash_estan_disponibles(self):
        nombre = self.carga(acciones=[{
            "momento": "tras_promover",
            "sql": "INSERT INTO auditoria VALUES ($carga, $fichero, $ejecucion_id)",
        }])
        self.csv()
        self.ejecutar(nombre)
        self.assertEqual(
            self.filas("SELECT que, valor FROM auditoria"), [("ventas", "ventas.csv")]
        )

    def test_un_parametro_llega_al_sql_y_no_solo_al_mapping(self):
        """Antes un parámetro solo podía alimentar una columna del mapping."""
        nombre = self.carga(
            parametros=[{"nombre": "tienda", "obligatorio": True}],
            acciones=[{
                "momento": "tras_promover",
                "sql": "INSERT INTO auditoria VALUES ('tienda', $p_tienda, $ejecucion_id)",
            }],
        )
        self.csv()
        motor_etl.ejecutar_carga(
            nombre, db_path=self.db_path, valores_parametros={"tienda": "Bilbao"}
        )
        self.assertEqual(self.escalar("SELECT valor FROM auditoria"), "Bilbao")


class PruebaTrasPromover(BaseVariables):
    def test_ve_el_resultado_de_la_escritura(self):
        nombre = self.carga(acciones=[{
            "momento": "tras_promover",
            "sql": "INSERT INTO auditoria VALUES ('promovidas', CAST($promovidas AS VARCHAR), $ejecucion_id)",
        }])
        self.csv()
        self.ejecutar(nombre)
        self.assertEqual(self.escalar("SELECT valor FROM auditoria"), "2")

    def test_ve_el_destino_ya_escrito(self):
        """Lo que `tras_validar` no puede: derivar del estado final sin
        reimplementar la singularidad."""
        nombre = self.carga(acciones=[{
            "momento": "tras_promover",
            "sql": "INSERT INTO auditoria SELECT 'total', CAST(sum(importe) AS VARCHAR), $ejecucion_id FROM destino",
        }])
        self.csv()
        self.ejecutar(nombre)
        self.assertEqual(self.escalar("SELECT valor FROM auditoria"), "30.0")

    def test_tras_validar_todavia_no_ve_las_filas_nuevas(self):
        """La diferencia entre los dos momentos, fijada."""
        nombre = self.carga(acciones=[{
            "momento": "tras_validar",
            "sql": "INSERT INTO auditoria SELECT 'antes', CAST(count(*) AS VARCHAR), $ejecucion_id FROM destino",
        }])
        self.csv()
        self.ejecutar(nombre)
        self.assertEqual(self.escalar("SELECT valor FROM auditoria"), "0")

    def test_si_la_accion_falla_se_revierte_tambien_la_promocion(self):
        """Misma transacción: o cuadra todo o no cuadra nada."""
        nombre = self.carga(acciones=[{
            "momento": "tras_promover", "sql": "INSERT INTO no_existe VALUES (1)",
        }])
        self.csv()
        with self.assertRaises(Exception):
            self.ejecutar(nombre)
        self.assertEqual(self.escalar("SELECT count(*) FROM destino"), 0)


class PruebaVariablesDeUsuario(BaseVariables):
    def test_cada_columna_del_select_es_una_variable(self):
        nombre = self.carga(
            variables=[{
                "momento": "tras_promover",
                "sql": "SELECT count(*) AS lineas, sum(importe) AS total FROM destino",
            }],
            acciones=[{
                "momento": "tras_promover",
                "sql": "INSERT INTO auditoria VALUES (CAST($v_lineas AS VARCHAR), CAST($v_total AS VARCHAR), $ejecucion_id)",
            }],
        )
        self.csv()
        self.ejecutar(nombre)
        self.assertEqual(self.filas("SELECT que, valor FROM auditoria"), [("2", "30.0")])

    def test_se_fija_al_capturarla_y_no_se_reevalua(self):
        """El valor es de cuando se capturó, aunque después cambie la tabla."""
        nombre = self.carga(
            variables=[{"momento": "tras_validar", "sql": "SELECT count(*) AS previas FROM destino"}],
            acciones=[{
                "momento": "tras_promover",
                "sql": "INSERT INTO auditoria VALUES ('previas', CAST($v_previas AS VARCHAR), $ejecucion_id)",
            }],
        )
        self.csv()
        self.ejecutar(nombre)
        # En tras_validar el destino estaba vacío; en tras_promover ya tiene 2.
        self.assertEqual(self.escalar("SELECT valor FROM auditoria"), "0")

    def test_cero_filas_falla_duro(self):
        nombre = self.carga(variables=[{
            "momento": "tras_validar", "sql": "SELECT 1 AS x WHERE false",
        }])
        self.csv()
        with self.assertRaises(motor_etl.VariableInvalidaError):
            self.ejecutar(nombre)

    def test_mas_de_una_fila_falla_duro(self):
        nombre = self.carga(variables=[{
            "momento": "tras_validar", "sql": "SELECT * FROM (VALUES (1), (2)) t(x)",
        }])
        self.csv()
        with self.assertRaises(motor_etl.VariableInvalidaError):
            self.ejecutar(nombre)


class PruebaValidacionDeVariables(PruebaConAlmacen):
    """Erratas cazadas al validar la definición, sin ejecutar nada."""

    def definicion(self, **extra):
        definicion = {
            "nombre": "v", "descripcion": DESCRIPCION, "patron": "*.csv",
            "formato": "csv", "delimitador": ";", "fila_cabecera": 1,
            "tabla_destino": "demo_venta",
            "mapping": [{"destino": "canal", "operaciones": [{"tipo": "const", "valor": "web"}]}],
        }
        definicion.update(extra)
        definicion.setdefault("carpeta", str(self.entrada_dir / "v"))
        return definicion

    def test_variable_inventada_se_rechaza(self):
        errores = cargas.validar(self.definicion(acciones=[
            {"momento": "antes", "sql": "SELECT $inventada"}
        ]))
        self.assertTrue(any("$inventada" in e for e in errores), errores)

    def test_promovidas_antes_de_promover_se_rechaza(self):
        """Existe, pero no todavía: el mensaje tiene que decir eso y no
        'variable desconocida', que mandaría a buscar una errata."""
        errores = cargas.validar(self.definicion(acciones=[
            {"momento": "tras_validar", "sql": "SELECT $promovidas"}
        ]))
        self.assertTrue(any("tras_promover" in e for e in errores), errores)

    def test_promovidas_en_tras_promover_se_acepta(self):
        errores = cargas.validar(self.definicion(acciones=[
            {"momento": "tras_promover", "sql": "SELECT $promovidas"}
        ]))
        self.assertEqual([e for e in errores if "promovidas" in e], [])

    def test_parametro_declarado_se_acepta_en_el_sql(self):
        errores = cargas.validar(self.definicion(
            parametros=[{"nombre": "tienda"}],
            acciones=[{"momento": "antes", "sql": "SELECT $p_tienda"}],
        ))
        self.assertEqual([e for e in errores if "p_tienda" in e], [])

    def test_celda_en_un_csv_se_rechaza(self):
        errores = cargas.validar(self.definicion(mapping=[
            {"destino": "canal", "operaciones": [{"tipo": "celda", "referencia": "B5"}]}
        ]))
        self.assertTrue(any("celda" in e and "excel" in e for e in errores), errores)


class PruebaOperacionCelda(PruebaConAlmacen):
    """El caso real: cabecera con metadatos arriba y la tabla más abajo."""

    def setUp(self):
        super().setUp()
        self.con.execute("""
            CREATE TABLE pedido (
                sucursal VARCHAR, mes VARCHAR, anio INTEGER,
                referencia VARCHAR, unidades INTEGER, extra_fields VARCHAR
            )
        """)
        self.escribir_catalogo(self.ficha_catalogo("pedido", {
            "sucursal": ("varchar", False), "mes": ("varchar", False),
            "anio": ("integer", False), "referencia": ("varchar", False),
            "unidades": ("integer", False), "extra_fields": ("varchar", False),
        }))
        self.ruta = self._escribir_xlsx()

    def _escribir_xlsx(self):
        import openpyxl

        carpeta = self.entrada_dir / "pedidos"
        carpeta.mkdir(parents=True, exist_ok=True)
        wb = openpyxl.Workbook()
        ws = wb.active
        ws["B5"] = "Bilbao"
        ws["B6"] = "03"        # texto a propósito: un mes con cero delante
        ws["B7"] = 2026
        ws["A10"], ws["B10"] = "Ref", "Unidades"
        for i, (ref, uds) in enumerate([("P-1", 3), ("P-2", 5)], start=11):
            ws[f"A{i}"], ws[f"B{i}"] = ref, uds
        ruta = carpeta / "pedidos_bilbao_marzo.xlsx"
        wb.save(ruta)
        return ruta

    def test_las_celdas_de_cabecera_se_reparten_por_todas_las_filas(self):
        nombre = self.escribir_carga({
            "nombre": "pedidos",
            "descripcion": DESCRIPCION,
            "carpeta": str(self.entrada_dir / "pedidos"),
            "patron": "*.xlsx",
            "formato": "excel",
            "fila_cabecera": 10,
            "tabla_destino": "pedido",
            "campos_singularidad": ["sucursal", "mes", "anio", "referencia"],
            "mapping": [
                {"origen": "Ref", "destino": "referencia", "operaciones": [{"tipo": "trim"}]},
                {"origen": "Unidades", "destino": "unidades",
                 "operaciones": [{"tipo": "cast", "tipo_destino": "integer"}]},
                {"destino": "sucursal", "operaciones": [{"tipo": "celda", "referencia": "B5"}]},
                {"destino": "mes", "operaciones": [{"tipo": "celda", "referencia": "B6"}]},
                {"destino": "anio", "operaciones": [
                    {"tipo": "celda", "referencia": "B7"},
                    {"tipo": "cast", "tipo_destino": "integer"},
                ]},
            ],
        })
        motor_etl.ejecutar_carga(nombre, db_path=self.db_path)
        self.assertEqual(
            self.filas("SELECT sucursal, mes, anio, referencia, unidades FROM pedido ORDER BY referencia"),
            [("Bilbao", "03", 2026, "P-1", 3), ("Bilbao", "03", 2026, "P-2", 5)],
        )

    def test_el_dry_run_ve_lo_mismo_que_la_carga(self):
        """Si el dry-run no leyera las celdas, esos campos saldrían vacíos
        en la previsualización y con valor al ejecutar."""
        nombre = self.escribir_carga({
            "nombre": "pedidos_seco",
            "descripcion": DESCRIPCION,
            "carpeta": str(self.entrada_dir / "pedidos"),
            "patron": "*.xlsx",
            "formato": "excel",
            "fila_cabecera": 10,
            "tabla_destino": "pedido",
            "mapping": [
                {"origen": "Ref", "destino": "referencia", "operaciones": [{"tipo": "trim"}]},
                {"destino": "sucursal", "operaciones": [{"tipo": "celda", "referencia": "B5"}]},
            ],
        })
        resultado = motor_etl.dry_run_carga(nombre, db_path=self.db_path)
        primera = resultado["ficheros"][0]["muestra_validas"][0]
        self.assertEqual(primera["sucursal"], "Bilbao")


class PruebaHistoricoAcumulativo(PruebaConAlmacen):
    """El patrón de la segunda tabla con singularidad propia, y su trampa.

    Alimentar un histórico desde `tras_promover` parece que se hace filtrando
    por el campo de negocio del fichero (`WHERE mes = $p_mes`). No: el destino
    acumula TODAS las sucursales, así que ese filtro se lleva también las filas
    de las demás que compartan mes, y el histórico duplica en cada carga. El
    filtro correcto es `ejecucion_id = $ejecucion_id`, que selecciona exacta y
    únicamente lo que acaba de escribir esta ejecución.
    """

    def setUp(self):
        super().setUp()
        self.con.execute("""
            CREATE TABLE pedido (
                sucursal VARCHAR, mes VARCHAR, referencia VARCHAR,
                extra_fields VARCHAR, ejecucion_id BIGINT
            )
        """)
        self.con.execute("""
            CREATE TABLE pedido_historico (
                sucursal VARCHAR, mes VARCHAR, referencia VARCHAR,
                extra_fields VARCHAR, ejecucion_id BIGINT
            )
        """)
        self.escribir_catalogo(self.ficha_catalogo("pedido", {
            "sucursal": ("varchar", False), "mes": ("varchar", False),
            "referencia": ("varchar", False), "extra_fields": ("varchar", False),
            "ejecucion_id": ("integer", False),
        }))

    def _carga(self, filtro_sql):
        return self.escribir_carga({
            "nombre": "pedidos",
            "descripcion": DESCRIPCION,
            "carpeta": str(self.entrada_dir / "pedidos"),
            "patron": "*.csv",
            "formato": "csv",
            "delimitador": ";",
            "fila_cabecera": 1,
            "tabla_destino": "pedido",
            "campos_singularidad": ["sucursal", "mes", "referencia"],
            "mapping": [
                {"origen": "Sucursal", "destino": "sucursal", "operaciones": [{"tipo": "trim"}]},
                {"origen": "Mes", "destino": "mes", "operaciones": [{"tipo": "trim"}]},
                {"origen": "Ref", "destino": "referencia", "operaciones": [{"tipo": "trim"}]},
            ],
            "acciones": [{
                "momento": "tras_promover",
                "sql": f"INSERT INTO pedido_historico SELECT * FROM pedido WHERE {filtro_sql}",
            }],
        })

    def _cargar(self, nombre, fichero, sucursal):
        self.escribir_csv(
            "pedidos", fichero,
            f"Sucursal;Mes;Ref\n{sucursal};03;P-1\n{sucursal};03;P-2\n",
        )
        motor_etl.ejecutar_carga(nombre, db_path=self.db_path)

    def test_filtrar_por_ejecucion_aisla_lo_que_escribio_esta_carga(self):
        nombre = self._carga("ejecucion_id = $ejecucion_id")
        self._cargar(nombre, "bilbao.csv", "Bilbao")
        self._cargar(nombre, "madrid.csv", "Madrid")

        self.assertEqual(self.escalar("SELECT count(*) FROM pedido"), 4)
        self.assertEqual(self.escalar("SELECT count(*) FROM pedido_historico"), 4)
        self.assertEqual(
            self.filas(
                "SELECT sucursal, count(*) FROM pedido_historico GROUP BY 1 ORDER BY 1"
            ),
            [("Bilbao", 2), ("Madrid", 2)],
        )

    def test_filtrar_por_un_campo_de_negocio_arrastra_las_otras_sucursales(self):
        """Por qué la prueba de arriba no es una obviedad."""
        nombre = self._carga("mes = '03'")
        self._cargar(nombre, "bilbao.csv", "Bilbao")
        self._cargar(nombre, "madrid.csv", "Madrid")

        # La segunda carga se lleva también las dos de Bilbao, que ya estaban.
        self.assertEqual(self.escalar("SELECT count(*) FROM pedido_historico"), 6)
        self.assertEqual(
            self.escalar("SELECT count(*) FROM pedido_historico WHERE sucursal = 'Bilbao'"), 4
        )
